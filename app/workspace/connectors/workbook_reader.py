from datetime import datetime
from html.parser import HTMLParser
from pathlib import Path
import re
from typing import Any, Protocol

import xlrd
from openpyxl import load_workbook


class WorkbookReadError(ValueError):
    pass


class WorkbookReaderAdapter(Protocol):
    def read(self, path: Path, worksheet: str | None, max_rows: int) -> dict[str, Any]: ...


class XlsxWorkbookReader:
    def read(self, path: Path, worksheet: str | None, max_rows: int) -> dict[str, Any]:
        stream = path.open("rb")
        try:
            workbook = load_workbook(stream, read_only=True, data_only=True)
        except Exception as exc:
            stream.close()
            raise WorkbookReadError("El archivo .xlsx está corrupto, protegido o no es un libro válido.") from exc
        try:
            sheets = workbook.sheetnames
            selected = worksheet or (sheets[0] if sheets else None)
            if not selected or selected not in sheets:
                raise WorkbookReadError("La hoja seleccionada no existe.")
            rows = workbook[selected].iter_rows(values_only=True)
            header = next(rows, None)
            values = []
            for number, row in enumerate(rows, start=2):
                if number > max_rows + 1:
                    raise WorkbookReadError(f"El archivo supera el límite de {max_rows} filas.")
                values.append((number, list(row)))
            return {"worksheets": sheets, "selected_worksheet": selected,
                    "header": list(header) if header else [], "rows": values}
        finally:
            workbook.close()
            stream.close()


class XlsWorkbookReader:
    def read(self, path: Path, worksheet: str | None, max_rows: int) -> dict[str, Any]:
        try:
            workbook = xlrd.open_workbook(path, on_demand=True)
        except Exception as exc:
            raise WorkbookReadError("El archivo .xls está corrupto, protegido o no es un libro válido.") from exc
        try:
            sheets = workbook.sheet_names()
            selected = worksheet or (sheets[0] if sheets else None)
            if not selected or selected not in sheets:
                raise WorkbookReadError("La hoja seleccionada no existe.")
            sheet = workbook.sheet_by_name(selected)
            if max(sheet.nrows - 1, 0) > max_rows:
                raise WorkbookReadError(f"El archivo supera el límite de {max_rows} filas.")
            header = sheet.row_values(0) if sheet.nrows else []
            rows = []
            for index in range(1, sheet.nrows):
                values = []
                for cell in sheet.row(index):
                    value = cell.value
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        value = datetime(*xlrd.xldate_as_tuple(value, workbook.datemode))
                    values.append(value)
                rows.append((index + 1, values))
            return {"worksheets": sheets, "selected_worksheet": selected,
                    "header": header, "rows": rows}
        finally:
            workbook.release_resources()


class _TableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.tables: list[dict[str, Any]] = []
        self._table: list[list[str]] | None = None
        self._row: list[str] | None = None
        self._cell: list[str] | None = None
        self._header_rows: set[int] = set()
        self._row_has_header = False
        self._orphan_headers: list[str] = []
        self._orphan_cell = False
        self._ignored_depth = 0

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        tag = tag.lower()
        if tag in {"script", "style", "form"}:
            self._ignored_depth += 1
        elif not self._ignored_depth and tag == "table":
            self._table = []
            self._header_rows = set()
            self._orphan_headers = []
        elif not self._ignored_depth and tag == "tr" and self._table is not None:
            self._row = []
            self._row_has_header = False
        elif not self._ignored_depth and tag in {"td", "th"} and self._row is not None:
            self._cell = []
            self._row_has_header = self._row_has_header or tag == "th"
            self._orphan_cell = False
        elif not self._ignored_depth and tag == "th" and self._table is not None:
            self._cell = []
            self._orphan_cell = True

    def handle_endtag(self, tag: str) -> None:
        tag = tag.lower()
        if tag in {"script", "style", "form"} and self._ignored_depth:
            self._ignored_depth -= 1
        elif not self._ignored_depth and tag in {"td", "th"} and self._cell is not None:
            value = "".join(self._cell).strip()
            if self._orphan_cell:
                self._orphan_headers.append(value)
            else:
                self._row.append(value)
            self._cell = None
            self._orphan_cell = False
        elif not self._ignored_depth and tag == "tr" and self._row is not None:
            if any(value for value in self._row):
                self._table.append(self._row)
                if self._row_has_header:
                    self._header_rows.add(len(self._table) - 1)
            self._row = None
        elif not self._ignored_depth and tag == "table" and self._table is not None:
            if self._orphan_headers:
                self._table.insert(0, self._orphan_headers)
                self._header_rows = {index + 1 for index in self._header_rows}
                self._header_rows.add(0)
            if self._table:
                self.tables.append({"rows": self._table,
                                    "header_rows": self._header_rows})
            self._table = None

    def handle_data(self, data: str) -> None:
        if not self._ignored_depth and self._cell is not None:
            self._cell.append(data)


class HtmlTableWorkbookReader:
    PRODUCT_HEADERS = {
        "sku", "numero parte", "número parte", "numero de parte",
        "part number", "internal sku", "referencia skf",
        "referencia fabricante", "referencia del fabricante",
        "manufacturer reference", "skf reference", "fob dd lista",
        "fob dd convenio", "precio convenio", "precio negociado",
        "agreement price", "contract price", "precio sugerido",
        "suggested price", "product line", "linea de producto",
        "línea de producto", "familia", "spc",
    }
    METADATA_KEYS = {
        "tipo": "agreement_type", "cliente": "customer_name",
        "fecha inicio": "start_date", "fecha fin": "end_date",
        "moneda": "currency", "currency": "currency",
        "proveedor": "supplier", "fabricante": "supplier",
    }

    def read(self, path: Path, worksheet: str | None, max_rows: int) -> dict[str, Any]:
        try:
            content = path.read_bytes().decode("utf-8-sig")
        except UnicodeDecodeError:
            try:
                content = path.read_bytes().decode("latin-1")
            except UnicodeDecodeError as exc:
                raise WorkbookReadError("El archivo exportado no contiene texto válido.") from exc
        parser = _TableParser()
        try:
            parser.feed(content)
            parser.close()
        except Exception as exc:
            raise WorkbookReadError("El archivo exportado está corrupto.") from exc
        usable = [table for table in parser.tables if table["rows"]]
        if not usable:
            raise WorkbookReadError("El archivo no contiene una tabla de datos utilizable.")
        names = [f"Tabla {index}" for index in range(1, len(usable) + 1)]
        metadata = self._metadata(usable)
        details = [self._describe(name, table, index) for index, (name, table)
                   in enumerate(zip(names, usable))]
        if worksheet:
            if worksheet not in names:
                raise WorkbookReadError("La tabla seleccionada no existe.")
            selected_index = names.index(worksheet)
            selected = worksheet
        else:
            selected_index = max(
                range(len(usable)),
                key=lambda index: self._product_score(usable[index]),
            )
            selected = names[selected_index]
        table = usable[selected_index]
        rows = table["rows"]
        header_index = self._header_index(table)
        if header_index is None:
            raise WorkbookReadError(
                "No fue posible identificar los encabezados de la tabla seleccionada."
            )
        if len(rows) - header_index - 1 > max_rows:
            raise WorkbookReadError(f"El archivo supera el límite de {max_rows} filas.")
        width = max(len(row) for row in rows)
        normalized = [row + [""] * (width - len(row)) for row in rows]
        return {
            "worksheets": names,
            "selected_worksheet": selected,
            "header": normalized[header_index],
            "rows": [
                (number, [self._cell_value(value) for value in row])
                for number, row in enumerate(
                    normalized[header_index + 1:], start=header_index + 2
                )
            ],
            "source_type": "html_xls",
            "source_table_index": selected_index,
            "detected_metadata": metadata,
            "worksheet_details": details,
            "detected_table_count": len(usable),
        }

    @classmethod
    def _normalized(cls, value: str) -> str:
        return " ".join(value.strip().lower().split())

    @classmethod
    def _header_index(cls, table: dict[str, Any]) -> int | None:
        if table["header_rows"]:
            return min(table["header_rows"])
        for index, row in enumerate(table["rows"][:10]):
            if sum(cls._normalized(value) in cls.PRODUCT_HEADERS for value in row) >= 1:
                return index
        return None

    @classmethod
    def _product_score(cls, table: dict[str, Any]) -> tuple[int, int, int]:
        rows = table["rows"]
        matches = max(
            (sum(cls._normalized(value) in cls.PRODUCT_HEADERS for value in row)
             for row in rows[:10]), default=0,
        )
        is_metadata = len(rows) <= 20 and max(map(len, rows), default=0) == 2
        return (matches - (100 if is_metadata else 0), len(rows),
                max(map(len, rows), default=0))

    @classmethod
    def _metadata(cls, tables: list[dict[str, Any]]) -> dict[str, Any]:
        result = {}
        for table in tables:
            rows = table["rows"]
            if len(rows) > 20 or max(map(len, rows), default=0) != 2:
                continue
            for row in rows:
                key = cls.METADATA_KEYS.get(cls._normalized(row[0]))
                if key and len(row) > 1:
                    result[key] = row[1].strip()
        return result

    @classmethod
    def _describe(cls, name: str, table: dict[str, Any], index: int) -> dict[str, Any]:
        rows = table["rows"]
        columns = max(map(len, rows), default=0)
        metadata = len(rows) <= 20 and columns == 2
        header_index = cls._header_index(table)
        product_rows = max(len(rows) - (header_index + 1 if header_index is not None else 0), 0)
        return {"name": name, "source_table_index": index,
                "row_count": product_rows, "column_count": columns,
                "kind": "metadata" if metadata else "products",
                "label": f"{name} — {len(rows) if metadata else product_rows} filas × {columns} columnas — "
                         f"{'Posibles metadatos' if metadata else 'Productos detectados'}"}

    @staticmethod
    def _cell_value(value: str) -> Any:
        cleaned = value.strip()
        decimal = re.fullmatch(r"([-+]?\d+)\s*,\s*(\d+)", cleaned)
        if decimal:
            return f"{decimal.group(1)}.{decimal.group(2)}"
        if re.fullmatch(r"[-+]?\d+(?:\.\d+)?", cleaned):
            return float(cleaned) if "." in cleaned else int(cleaned)
        return cleaned


class WorkbookReader:
    XLSX_SIGNATURE = b"PK\x03\x04"
    XLS_SIGNATURE = bytes.fromhex("D0CF11E0A1B11AE1")

    @classmethod
    def read(cls, path: Path, worksheet: str | None, max_rows: int) -> dict[str, Any]:
        reader = cls._detect(path)
        return reader.read(path, worksheet, max_rows)

    @classmethod
    def _detect(cls, path: Path) -> WorkbookReaderAdapter:
        prefix = path.read_bytes()[:4096]
        if prefix.startswith(cls.XLSX_SIGNATURE):
            return XlsxWorkbookReader()
        if prefix.startswith(cls.XLS_SIGNATURE):
            return XlsWorkbookReader()
        text = prefix.decode("latin-1", errors="ignore").lstrip().lower()
        if text.startswith(("<html", "<!doctype html", "<table")) or "<table" in text:
            return HtmlTableWorkbookReader()
        raise WorkbookReadError(
            "El archivo no contiene un libro Excel ni una exportación ERP válida."
        )
