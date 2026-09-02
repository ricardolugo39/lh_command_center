import sqlite3
import io

import pytest

from app.database.migrations import upgrade
from app import create_app
from app.workspace.stock_planning.service import StockPlanningFoundationService
from app.workspace.stock_planning.repository import StockPlanningRepository
from app.workspace.stock_planning.forecasting import StockForecastEngine
from app.workspace.stock_planning.decisions import StockPlanningDecisionService
from app.workspace.stock_planning.exports import StockPlanningExportService
from app.workspace.stock_planning.replenishment import StockReplenishmentService


@pytest.fixture
def stock_database(tmp_path, monkeypatch):
    database = tmp_path / "stock-planning.db"
    monkeypatch.setattr("app.database.connection.DB_PATH", database)
    upgrade()
    with sqlite3.connect(database) as connection:
        connection.executescript(
            """CREATE TABLE raw_sales (
                fecha TEXT, idproducto TEXT, nombreproducto TEXT,
                idbodega INTEGER, cantidad REAL, sufijo TEXT
            );
            INSERT INTO raw_sales VALUES
                ('2025-07-01','SOLD-ONLY','Sold only',1,2,'VENDORX');
            INSERT INTO inventario_snapshot (
                fecha_snapshot,idbodega,nombre_bodega,idproducto,
                nombreproducto,unidades,unidades_disponible,
                unidades_reservado,unidades_remisionado,unidades_transito,
                marca_codigo,marca_nombre,archivo_origen
            ) VALUES
                ('2025-07-15','1','Bogotá','STOCKED','Stocked',10,7,2,1,3,
                 'VX','Vendor X','inventory.xlsx'),
                ('2025-07-15','900','Tránsito','NOT-A-PLANNING-BRANCH','Other',5,5,0,0,0,
                 'VX','Vendor X','inventory.xlsx');
            """
        )
    return database


def _configure():
    StockPlanningFoundationService.register_branch(
        branch_code="1", branch_name="Bogotá", is_primary_receipt=1,
    )
    StockPlanningFoundationService.register_branch(
        branch_code="50", branch_name="Cali",
    )
    profile_id = StockPlanningFoundationService.create_vendor_profile(
        vendor_name="Vendor X", profile_code="vendor_x",
        inventory_brand_codes=["VX"], sales_suffixes=["VENDORX"],
    )
    StockPlanningFoundationService.register_catalog_product(
        profile_id, internal_sku="ZERO-STOCK", vendor_sku="ZERO",
        product_name="Zero inventory catalogue product",
    )
    return profile_id


def _insert_fob_price(database, sku, fob, *, suffix="VX", nit="9001"):
    with sqlite3.connect(database) as connection:
        cursor = connection.execute(
            """INSERT INTO erp_import_executions (
                import_type,original_filename,stored_file_path,file_hash,
                schema_version,status,executed_by
            ) VALUES ('fob_prices','fob.xlsx','/tmp/fob.xlsx',?,
                      'erp-v1','completed','tester')""",
            (f"hash-{sku}-{fob}",),
        )
        connection.execute(
            """INSERT INTO erp_fob_price_history (
                import_execution_id,idproducto,prefijo,sufijo,idfam2,
                fob_usd,lista1_cop,nit
            ) VALUES (?,?,?,?,?,?,?,?)""",
            (cursor.lastrowid, sku, sku, suffix, "10", fob, 100000, nit),
        )


def test_thomson_profile_is_enabled_with_erp_aliases(stock_database):
    with sqlite3.connect(stock_database) as connection:
        profile = connection.execute(
            """SELECT vendor_name,inventory_brand_codes_json,sales_suffixes_json
            FROM stock_planning_vendor_profiles WHERE profile_code='Thomson'"""
        ).fetchone()

    assert profile == ("Thomson", '["THO"]', '["THO"]')


def test_replenishment_profiles_and_cali_frequency_policy(stock_database):
    profiles = {
        item["profile_code"] for item in StockPlanningRepository.list_vendor_profiles()
    }
    assert {"SKF", "FAG", "NTN", "NQK", "KMK"}.issubset(profiles)
    assert StockReplenishmentService._automatic(
        {"active_months_24": 6, "sales_12": 8, "review_reasons": []}
    )
    assert StockReplenishmentService._eligible(
        {"active_months_24": 3, "sales_12": 2}
    )
    assert not StockReplenishmentService._automatic(
        {"active_months_24": 5, "sales_12": 8, "review_reasons": []}
    )
    assert not StockReplenishmentService._automatic(
        {"active_months_24": 8, "sales_12": 0, "review_reasons": []}
    )
    assert not StockReplenishmentService._automatic({
        "active_months_24": 8, "sales_12": 8,
        "review_reasons": ["Inventario utilizable negativo"],
    })
    assumptions = StockReplenishmentService._assumptions(90)
    assert round(
        assumptions["coverage_months"] * 30.4375
        + assumptions["cali_transfer_days"]
    ) == 90
    assert StockReplenishmentService._to_cali(
        {"from_branch": "1", "to_branch": "50"}
    )
    assert not StockReplenishmentService._to_cali(
        {"from_branch": "50", "to_branch": "1"}
    )


def test_replenishment_skips_incomplete_inventory_without_duplicate_run(stock_database):
    from datetime import date

    first = StockReplenishmentService.run_due(
        today=date(2025, 7, 15), triggered_by="test", force=True
    )
    second = StockReplenishmentService.run_due(
        today=date(2025, 7, 15), triggered_by="test", force=True
    )

    assert len(first) == 5
    assert all(item["status"] == "skipped" for item in first)
    assert len(second) == 5
    with sqlite3.connect(stock_database) as connection:
        assert connection.execute(
            "SELECT COUNT(*) FROM stock_planning_replenishment_runs"
        ).fetchone()[0] == 5
        assert connection.execute(
            "SELECT COUNT(*) FROM stock_planning_notifications"
        ).fetchone()[0] == 5


def test_snapshot_unites_catalog_sales_inventory_and_branches(stock_database):
    profile_id = _configure()
    snapshot = StockPlanningFoundationService.create_snapshot(
        profile_id=profile_id, as_of_date="2025-07-15", created_by="tester",
    )

    assert snapshot.product_count == 3
    assert snapshot.inventory_row_count == 9
    assert snapshot.inventory_snapshot_date == "2025-07-15"
    assert snapshot.sales_through_date == "2025-07-01"
    with sqlite3.connect(stock_database) as connection:
        connection.row_factory = sqlite3.Row
        products = connection.execute(
            """SELECT internal_sku,is_catalog_product,has_sales_history,
                has_inventory_history FROM stock_planning_snapshot_products
            WHERE snapshot_id=? ORDER BY internal_sku""", (snapshot.snapshot_id,)
        ).fetchall()
        positions = connection.execute(
            """SELECT branch_code,internal_sku,on_hand,usable
            FROM stock_planning_snapshot_inventory WHERE snapshot_id=?
            ORDER BY branch_code,internal_sku""", (snapshot.snapshot_id,)
        ).fetchall()

    assert [row["internal_sku"] for row in products] == [
        "SOLD-ONLY", "STOCKED", "ZERO-STOCK",
    ]
    assert dict(products[0])["has_sales_history"] == 1
    assert dict(products[1])["has_inventory_history"] == 1
    assert dict(products[2])["is_catalog_product"] == 1
    assert {(row["branch_code"], row["internal_sku"]) for row in positions} == {
        (branch, sku)
        for branch in ("1", "16", "50")
        for sku in ("SOLD-ONLY", "STOCKED", "ZERO-STOCK")
    }
    bogota_stock = next(
        row for row in positions
        if row["branch_code"] == "1" and row["internal_sku"] == "STOCKED"
    )
    assert bogota_stock["on_hand"] == 10
    assert bogota_stock["usable"] == 7


def test_product_sales_movements_are_frozen_and_consolidate_bogota(stock_database):
    with sqlite3.connect(stock_database) as connection:
        connection.execute("ALTER TABLE raw_sales ADD COLUMN razonsocial TEXT")
        connection.execute("ALTER TABLE raw_sales ADD COLUMN neto REAL")
        connection.execute("ALTER TABLE raw_sales ADD COLUMN nombrebodega TEXT")
        connection.execute(
            """UPDATE raw_sales SET razonsocial='Cliente Uno',neto=200000,
                nombrebodega='Principal' WHERE idproducto='SOLD-ONLY'"""
        )
        connection.execute(
            """INSERT INTO raw_sales (
                fecha,idproducto,nombreproducto,idbodega,cantidad,sufijo,
                razonsocial,neto,nombrebodega
            ) VALUES ('2025-06-15','SOLD-ONLY','Sold only',16,3,'VENDORX',
                      'Cliente Dos',300000,'KR 68')"""
        )
    profile_id = _configure()
    snapshot = StockPlanningFoundationService.create_snapshot(
        profile_id=profile_id, as_of_date="2025-07-15", created_by="tester",
    )
    # A later source correction must not alter the frozen evidence.
    with sqlite3.connect(stock_database) as connection:
        connection.execute(
            """INSERT INTO raw_sales (
                fecha,idproducto,nombreproducto,idbodega,cantidad,sufijo,
                razonsocial,neto,nombrebodega
            ) VALUES ('2025-07-10','SOLD-ONLY','Sold only',1,99,'VENDORX',
                      'Cliente Tardío',9900000,'Principal')"""
        )

    evidence = StockPlanningRepository.product_sales_movements(
        snapshot.snapshot_id, "SOLD-ONLY", "1", 36
    )

    assert evidence["summary"] == {
        "movement_count": 2, "net_units": 5.0,
        "customer_count": 2, "net_value_cop": 500000.0,
    }
    assert {row["warehouse_code"] for row in evidence["rows"]} == {"1", "16"}
    assert all("nit" not in row for row in evidence["rows"])


def test_snapshot_freezes_latest_fob_and_order_total_uses_final_quantity(
    stock_database, monkeypatch,
):
    profile_id = _configure()
    _insert_fob_price(stock_database, "STOCKED", 12.5)
    snapshot = StockPlanningFoundationService.create_snapshot(
        profile_id=profile_id, as_of_date="2025-07-15", created_by="tester",
    )
    _insert_fob_price(stock_database, "STOCKED", 15.0)

    forecast = {
        "rows": [
            {**_forecast_row("STOCKED", "1", 4), "requires_review": True},
            _forecast_row("SOLD-ONLY", "50", 2),
        ],
        "transfers": [],
    }
    monkeypatch.setattr(StockForecastEngine, "analyze", lambda _: forecast)
    StockPlanningDecisionService.save_purchase(
        snapshot.snapshot_id, "STOCKED", "1", 3, "buyer@example.com"
    )
    presented = StockPlanningDecisionService.present(
        snapshot.snapshot_id, forecast
    )

    stocked = presented["rows"][0]
    assert stocked["fob_usd"] == 12.5
    assert stocked["total_fob_usd"] == 37.5
    assert presented["total_fob_usd"] == 37.5
    assert presented["missing_fob_skus"] == ["SOLD-ONLY"]
    assert presented["missing_fob_count"] == 1

    newer = StockPlanningFoundationService.create_snapshot(
        profile_id=profile_id, as_of_date="2025-07-15", created_by="tester",
    )
    with sqlite3.connect(stock_database) as connection:
        prices = connection.execute(
            """SELECT snapshot_id,fob_usd
            FROM stock_planning_snapshot_fob_prices
            WHERE internal_sku='STOCKED' ORDER BY snapshot_id"""
        ).fetchall()
    assert prices == [
        (snapshot.snapshot_id, 12.5),
        (newer.snapshot_id, 15.0),
    ]


def test_dated_and_undated_transit_are_not_combined(stock_database):
    profile_id = _configure()
    StockPlanningFoundationService.register_transit_supply(
        profile_id, branch_code="50", internal_sku="ZERO-STOCK",
        quantity=4, expected_date="2025-10-21", purchase_order_reference="PO-1",
    )
    snapshot = StockPlanningFoundationService.create_snapshot(
        profile_id=profile_id, as_of_date="2025-07-15", created_by="tester",
    )

    with sqlite3.connect(stock_database) as connection:
        position = connection.execute(
            """SELECT undated_transit,dated_transit
            FROM stock_planning_snapshot_inventory
            WHERE snapshot_id=? AND branch_code='50' AND internal_sku='ZERO-STOCK'""",
            (snapshot.snapshot_id,),
        ).fetchone()
        issues = connection.execute(
            """SELECT issue_code,branch_code,internal_sku
            FROM stock_planning_snapshot_issues WHERE snapshot_id=?""",
            (snapshot.snapshot_id,),
        ).fetchall()

    assert position == (0, 4)
    assert ("UNDATED_TRANSIT", "1", "STOCKED") in issues


def test_frozen_snapshot_rejects_mutation(stock_database):
    profile_id = _configure()
    snapshot = StockPlanningFoundationService.create_snapshot(
        profile_id=profile_id, as_of_date="2025-07-15", created_by="tester",
    )

    with sqlite3.connect(stock_database) as connection:
        with pytest.raises(sqlite3.IntegrityError, match="immutable"):
            connection.execute(
                "UPDATE stock_planning_snapshots SET as_of_date='2025-07-16' WHERE id=?",
                (snapshot.snapshot_id,),
            )


def test_new_master_data_does_not_modify_prior_snapshot(stock_database):
    profile_id = _configure()
    first = StockPlanningFoundationService.create_snapshot(
        profile_id=profile_id, as_of_date="2025-07-15", created_by="tester",
    )
    StockPlanningFoundationService.register_catalog_product(
        profile_id, internal_sku="LATER", vendor_sku="LATER",
    )
    second = StockPlanningFoundationService.create_snapshot(
        profile_id=profile_id, as_of_date="2025-07-15", created_by="tester",
    )

    assert first.product_count == 3
    assert second.product_count == 4
    with sqlite3.connect(stock_database) as connection:
        first_count = connection.execute(
            "SELECT COUNT(*) FROM stock_planning_snapshot_products WHERE snapshot_id=?",
            (first.snapshot_id,),
        ).fetchone()[0]
    assert first_count == 3


def test_family_and_transformation_products_enter_complete_universe(stock_database):
    profile_id = _configure()
    family_id = StockPlanningFoundationService.register_family(
        profile_id, family_code="LINEAR_25", family_name="Linear size 25",
    )
    StockPlanningFoundationService.register_family_member(
        family_id, internal_sku="BLOCK-25", relationship_role="block",
        confidence=0.9,
    )
    StockPlanningFoundationService.register_transformation(
        profile_id, transformation_code="RAIL_25_3M",
        transformation_type="length_cut", purchase_sku="RAIL-25-3000",
        purchase_quantity=1, waste_rate=0.05,
        inputs=[{
            "sales_sku": "RAIL-25-1000", "sales_quantity": 1,
            "normalized_consumption": 1000,
        }],
    )
    snapshot = StockPlanningFoundationService.create_snapshot(
        profile_id=profile_id, as_of_date="2025-07-15", created_by="tester",
    )

    with sqlite3.connect(stock_database) as connection:
        skus = {
            row[0] for row in connection.execute(
                """SELECT internal_sku FROM stock_planning_snapshot_products
                WHERE snapshot_id=?""", (snapshot.snapshot_id,),
            ).fetchall()
        }
    assert {"BLOCK-25", "RAIL-25-1000", "RAIL-25-3000"}.issubset(skus)


def test_administration_ui_configures_and_reviews_snapshot(stock_database):
    application = create_app({"TESTING": True, "TEST_AUTH_BYPASS": True})
    client = application.test_client()

    landing = client.get("/stock-planning/")
    assert landing.status_code == 200
    assert "Planeación de inventario" in landing.get_data(as_text=True)
    vendor = client.post("/stock-planning/vendors", data={
        "vendor_name": "Vendor X", "profile_code": "vendor_x",
        "inventory_brand_codes": "VX", "sales_suffixes": "VENDORX",
    }, follow_redirects=True)
    assert vendor.status_code == 200
    assert "Proveedor configurado" in vendor.get_data(as_text=True)
    operational_html = vendor.get_data(as_text=True)
    assert "Analizar pedido de hoy" in operational_html
    assert "Configurar proveedor" not in operational_html
    assert "Configurar bodegas" not in operational_html
    assert "Catálogo de compra" not in operational_html
    assert "Tránsito con fecha" not in operational_html
    assert "Familias de productos" not in operational_html
    assert "Transformación compra" not in operational_html

    with sqlite3.connect(stock_database) as connection:
        profile_id = connection.execute(
            "SELECT id FROM stock_planning_vendor_profiles WHERE profile_code='VENDOR_X'"
        ).fetchone()[0]
    client.post("/stock-planning/branches", data={
        "profile_id": profile_id, "branch_code": "1",
        "branch_name": "Bogotá", "is_primary_receipt": "on",
    })
    created = client.post("/stock-planning/snapshots", data={
        "profile_id": profile_id,
        "manufacturing_days": "60",
        "international_shipping_days": "30",
        "receiving_days": "5",
        "cali_transfer_days": "7",
        "coverage_months": "6",
    }, follow_redirects=True)
    html = created.get_data(as_text=True)
    assert created.status_code == 200
    assert "Análisis guardado" in html
    assert "STOCKED" in html
    assert "Tránsito sin fecha" in html
    with sqlite3.connect(stock_database) as connection:
        inputs = connection.execute(
            """SELECT manufacturing_days,international_shipping_days,
                receiving_days,cali_transfer_days,coverage_months
            FROM stock_planning_analysis_inputs"""
        ).fetchone()
    assert inputs == (60, 30, 5, 7, 6)


def test_selected_vendor_is_remembered_and_test_analysis_can_be_archived(
    stock_database,
):
    profile_id = _configure()
    snapshot = StockPlanningFoundationService.create_snapshot(
        profile_id=profile_id, as_of_date="2025-07-15", created_by="tester",
        assumptions={"manufacturing_days": 30, "international_shipping_days": 30,
                     "receiving_days": 5, "cali_transfer_days": 3,
                     "coverage_months": 6},
    )
    application = create_app({"TESTING": True, "TEST_AUTH_BYPASS": True})
    client = application.test_client()

    selected = client.get(f"/stock-planning/?profile_id={profile_id}")
    remembered = client.get("/stock-planning/")
    assert f'<option value="{profile_id}" selected>Vendor X</option>' in (
        selected.get_data(as_text=True)
    )
    assert f'<option value="{profile_id}" selected>Vendor X</option>' in (
        remembered.get_data(as_text=True)
    )

    archived = client.post(
        f"/stock-planning/snapshots/{snapshot.snapshot_id}/archive",
        follow_redirects=True,
    )
    assert archived.status_code == 200
    assert "Análisis eliminado de la lista" in archived.get_data(as_text=True)
    assert client.get(
        f"/stock-planning/snapshots/{snapshot.snapshot_id}"
    ).status_code == 404
    with sqlite3.connect(stock_database) as connection:
        row = connection.execute(
            """SELECT archived_at,archived_by FROM stock_planning_snapshots
            WHERE id=?""", (snapshot.snapshot_id,),
        ).fetchone()
        product_count = connection.execute(
            """SELECT COUNT(*) FROM stock_planning_snapshot_products
            WHERE snapshot_id=?""", (snapshot.snapshot_id,),
        ).fetchone()[0]
    assert row[0] is not None
    assert row[1]
    assert product_count > 0


def test_analysis_history_shows_all_brands_regardless_of_selected_vendor(
    stock_database,
):
    profile_id = _configure()
    StockPlanningFoundationService.create_snapshot(
        profile_id=profile_id, as_of_date="2025-07-15", created_by="tester",
    )
    thomson_id = next(
        profile["id"] for profile in StockPlanningRepository.list_vendor_profiles()
        if profile["vendor_name"] == "Thomson"
    )
    StockPlanningFoundationService.create_snapshot(
        profile_id=thomson_id, as_of_date="2025-07-15", created_by="tester",
    )

    page = StockPlanningFoundationService.dashboard(profile_id)

    assert {snapshot["vendor_name"] for snapshot in page["snapshots"]} == {
        "Vendor X", "Thomson",
    }


def _forecast_row(sku, branch, order, usable=0, target=0):
    return {
        "sku": sku, "branch": branch, "recommended_order": order,
        "usable": usable, "transit": 0, "target_stock": target,
        "review_reasons": [], "requires_review": False,
        "abc": "B", "xyz": "Y", "model": "Promedio 12",
    }


def test_transfer_uses_only_donor_surplus_before_buying():
    bogota = _forecast_row("BLOCK-25", "1", 0, usable=20, target=8)
    cali = _forecast_row("BLOCK-25", "50", 10, usable=0, target=10)

    transfers = StockForecastEngine._apply_transfers([bogota, cali])

    assert transfers == [{
        "sku": "BLOCK-25", "from_branch": "1", "to_branch": "50",
        "quantity": 10, "avoided_purchase": 10,
        "bogota_inventory": 20.0, "cali_inventory": 0.0,
        "internal_move_16_to_1": 0,
    }]
    assert cali["recommended_order"] == 0
    assert bogota["transfer_out"] == 10
    assert bogota["usable"] - bogota["transfer_out"] >= bogota["target_stock"]


def test_warehouse_16_is_consolidated_into_bogota_and_retained_as_evidence():
    positions = StockForecastEngine._planning_positions([
        {
            "branch_code": "1", "branch_name": "Bogotá", "internal_sku": "A",
            "on_hand": 4, "reserved": 0, "remitted": 0, "usable": 4,
            "undated_transit": 1, "dated_transit": 0, "average_cost": 10,
        },
        {
            "branch_code": "16", "branch_name": "KR 68", "internal_sku": "A",
            "on_hand": 7, "reserved": 1, "remitted": 0, "usable": 6,
            "undated_transit": 0, "dated_transit": 0, "average_cost": 10,
        },
        {
            "branch_code": "50", "branch_name": "Cali", "internal_sku": "A",
            "on_hand": 2, "reserved": 0, "remitted": 0, "usable": 2,
            "undated_transit": 0, "dated_transit": 0, "average_cost": 10,
        },
    ])

    by_branch = {row["branch_code"]: row for row in positions}
    assert set(by_branch) == {"1", "50"}
    assert by_branch["1"]["usable"] == 10
    assert by_branch["1"]["warehouse_1_usable"] == 4
    assert by_branch["1"]["warehouse_16_usable"] == 6

    bogota = _forecast_row("A", "1", 0, usable=10, target=2)
    bogota["warehouse_16_usable"] = 6
    cali = _forecast_row("A", "50", 5, usable=0, target=5)
    transfer = StockForecastEngine._apply_transfers([bogota, cali])[0]
    assert transfer["from_branch"] == "1"
    assert transfer["to_branch"] == "50"
    assert transfer["internal_move_16_to_1"] == 5


def test_rails_and_ball_screws_consolidate_into_three_meter_bars():
    rows = [
        _forecast_row("HSR 25-1000LTHK", "1", 1),
        _forecast_row("HSR 25-2000LTHK", "1", 1),
        _forecast_row("HSR 25-3000LTHK", "1", 0),
        _forecast_row("TS 2510+1000LTHK", "50", 2),
        _forecast_row("TS 2510+2000LTHK", "50", 1),
        _forecast_row("TS 2510+3000LTHK", "50", 0),
    ]

    transformed, evidence = StockForecastEngine._apply_length_transformations(rows)
    by_sku_branch = {(row["sku"], row["branch"]): row for row in transformed}

    rail = by_sku_branch[("HSR 25-3000LTHK", "1")]
    screw = by_sku_branch[("TS 2510+3000LTHK", "50")]
    assert rail["recommended_order"] == 1
    assert rail["required_length_mm"] == 3000
    assert screw["recommended_order"] == 2
    assert screw["required_length_mm"] == 4000
    assert len(evidence) == 2
    assert by_sku_branch[("HSR 25-1000LTHK", "1")]["recommended_order"] == 0
    assert by_sku_branch[("TS 2510+2000LTHK", "50")]["recommended_order"] == 0


def test_thomson_standard_lengths_remain_independent():
    rows = [
        _forecast_row("W 20 H6/1000THO", "1", 3),
        _forecast_row("W 20 H6/2000THO", "1", 2),
        _forecast_row("W 20 H6/3000THO", "1", 1),
    ]

    transformed, evidence = StockForecastEngine._apply_length_transformations(
        rows, "Thomson"
    )

    assert [row["recommended_order"] for row in transformed] == [3, 2, 1]
    assert evidence == []


def test_thomson_special_lengths_use_best_available_standard_bar_mix():
    rows = [
        _forecast_row("W 40 H6/1000THO", "50", 0),
        _forecast_row("W 40 H6/2000THO", "50", 0),
        _forecast_row("W 40 H6/3000THO", "50", 0),
        _forecast_row("W 40 H6/1255THO", "50", 2),
    ]

    transformed, evidence = StockForecastEngine._apply_length_transformations(
        rows, "thomson"
    )
    by_sku = {row["sku"]: row for row in transformed}

    assert by_sku["W 40 H6/1255THO"]["recommended_order"] == 0
    assert by_sku["W 40 H6/3000THO"]["recommended_order"] == 1
    assert by_sku["W 40 H6/3000THO"]["special_cut_bars"] == 1
    assert evidence[0]["required_mm"] == 2510
    assert evidence[0]["waste_mm"] == 490


def test_thomson_four_meter_shaft_requires_review():
    row = _forecast_row("W 20 H6/4000THO", "1", 1)

    StockForecastEngine._apply_length_transformations([row], "thomson")

    assert row["recommended_order"] == 1
    assert row["requires_review"] is True
    assert "Eje Thomson de 4 metros" in row["review_reasons"]


def test_purchase_and_transfer_decisions_preserve_suggestion(stock_database, monkeypatch):
    profile_id = _configure()
    snapshot = StockPlanningFoundationService.create_snapshot(
        profile_id=profile_id, as_of_date="2025-07-15", created_by="tester",
        assumptions={"manufacturing_days": 30, "international_shipping_days": 30,
                     "receiving_days": 5, "cali_transfer_days": 3,
                     "coverage_months": 6},
    )
    forecast = {
        "rows": [{**_forecast_row("BLOCK-25", "1", 10),
                  "requires_review": True}],
        "transfers": [{"sku": "BLOCK-30", "from_branch": "50",
                       "to_branch": "1", "quantity": 4, "avoided_purchase": 4}],
    }
    monkeypatch.setattr(StockForecastEngine, "analyze", lambda _: forecast)

    StockPlanningDecisionService.save_purchase(
        snapshot.snapshot_id, "BLOCK-25", "1", 8, "buyer@example.com", "Ajustado"
    )
    StockPlanningDecisionService.save_transfer(
        snapshot.snapshot_id, "BLOCK-30", "50", "1", 3, "buyer@example.com"
    )
    presented = StockPlanningDecisionService.present(snapshot.snapshot_id, forecast)

    purchase = presented["rows"][0]
    transfer = presented["transfers"][0]
    assert purchase["recommended_order"] == 10
    assert purchase["final_quantity"] == 8
    assert purchase["decision"]["decision_status"] == "changed"
    assert transfer["quantity"] == 4
    assert transfer["final_quantity"] == 3
    assert presented["purchase_export_ready"] is True
    assert presented["transfer_export_ready"] is True
    with sqlite3.connect(stock_database) as connection:
        history = connection.execute(
            "SELECT COUNT(*) FROM stock_planning_decision_history"
        ).fetchone()[0]
    assert history == 2


def test_decision_redirects_return_to_the_edited_section(stock_database, monkeypatch):
    profile_id = _configure()
    snapshot = StockPlanningFoundationService.create_snapshot(
        profile_id=profile_id, as_of_date="2025-07-15", created_by="tester",
        assumptions={"manufacturing_days": 30, "international_shipping_days": 30,
                     "receiving_days": 5, "cali_transfer_days": 3,
                     "coverage_months": 6},
    )
    forecast = {
        "rows": [{**_forecast_row("BLOCK-25", "1", 10),
                  "requires_review": True}],
        "transfers": [{"sku": "BLOCK-30", "from_branch": "50",
                       "to_branch": "1", "quantity": 4, "avoided_purchase": 4}],
    }
    monkeypatch.setattr(StockForecastEngine, "analyze", lambda _: forecast)
    application = create_app({"TESTING": True, "TEST_AUTH_BYPASS": True})
    client = application.test_client()

    purchase = client.post(
        f"/stock-planning/snapshots/{snapshot.snapshot_id}/purchase-decisions",
        data={"sku": "BLOCK-25", "branch": "1", "quantity": "8"},
    )
    product = client.post(
        f"/stock-planning/snapshots/{snapshot.snapshot_id}/purchase-decisions",
        data={"sku": "BLOCK-25", "branch": "1", "quantity": "8",
              "return_to": "product"},
    )
    transfer = client.post(
        f"/stock-planning/snapshots/{snapshot.snapshot_id}/transfer-decisions",
        data={"sku": "BLOCK-30", "from_branch": "50", "to_branch": "1",
              "quantity": "3"},
    )

    assert purchase.location.endswith("#purchase-decisions")
    assert product.location.endswith("#purchase-decision")
    assert transfer.location.endswith("#transfer-decisions")


def test_excel_exports_include_branch_and_approved_quantities(monkeypatch):
    from openpyxl import load_workbook

    page = {"snapshot": {"vendor_name": "Vendor X", "snapshot_key": "SP-TEST",
                         "as_of_date": "2025-07-15"},
            "products": [{"internal_sku": "BLOCK-25", "vendor_sku": "V-BLOCK-25"}]}
    purchase = {**_forecast_row("BLOCK-25", "1", 10), "final_quantity": 8,
                "on_hand": 5, "decision": {"decision_status": "changed"},
                "requires_review": True, "fob_usd": 12.5,
                "total_fob_usd": 100.0}
    forecast = {
        "rows": [purchase], "transfers": [
            {
                "sku": "BLOCK-30", "from_branch": "50", "to_branch": "1",
                "quantity": 4, "final_quantity": 3, "avoided_purchase": 4,
                "decision": {"decision_status": "changed"},
            },
            {
                "sku": "BLOCK-40", "from_branch": "1", "to_branch": "50",
                "quantity": 2, "final_quantity": 2, "avoided_purchase": 2,
                "internal_move_16_to_1": 2,
                "decision": {"decision_status": "approved"},
            },
        ], "purchase_export_ready": True, "transfer_export_ready": True,
    }
    monkeypatch.setattr(StockPlanningExportService, "_data", lambda _: (page, forecast))

    purchase_stream, _ = StockPlanningExportService.purchase_order(1)
    transfer_stream, _ = StockPlanningExportService.transfers(1)
    purchase_sheet = load_workbook(io.BytesIO(purchase_stream.read())).active
    transfer_book = load_workbook(io.BytesIO(transfer_stream.read()))
    cali_sheet = transfer_book["Despachos desde Cali"]
    bogota_sheet = transfer_book["Despachos desde Bogotá"]

    assert [cell.value for cell in purchase_sheet[5]][:5] == [
        "Bodega", "Código bodega", "Referencia proveedor", "Referencia interna",
        "Cantidad aprobada"
    ]
    assert [cell.value for cell in purchase_sheet[6]][:5] == [
        "Bogotá", "1", "V-BLOCK-25", "BLOCK-25", 8
    ]
    purchase_headers = [cell.value for cell in purchase_sheet[5]]
    assert purchase_sheet.cell(
        6, purchase_headers.index("FOB unitario USD") + 1
    ).value == 12.5
    assert purchase_sheet.cell(
        6, purchase_headers.index("Total FOB USD") + 1
    ).value == 100.0
    expected_transfer_headers = [
        "Referencia", "Desde", "Código origen", "Hacia",
        "Código destino", "Cantidad aprobada", "Movimiento previo",
    ]
    assert [cell.value for cell in cali_sheet[5]] == expected_transfer_headers
    assert [cell.value for cell in bogota_sheet[5]] == expected_transfer_headers
    assert [cell.value for cell in cali_sheet[6]][:6] == [
        "BLOCK-30", "Cali", "50", "Bogotá", "1", 3
    ]
    assert [cell.value for cell in bogota_sheet[6]][:6] == [
        "BLOCK-40", "Bogotá", "1", "Cali", "50", 2
    ]
    assert bogota_sheet[6][6].value == (
        "Mover primero 2 unidad(es) de bodega 16 (KR 68) a bodega 1"
    )
    removed = {"Cantidad sugerida", "Compra evitada", "Estado decisión"}
    assert removed.isdisjoint(cell.value for cell in cali_sheet[5])
    assert removed.isdisjoint(cell.value for cell in bogota_sheet[5])


def test_replenishment_uncovered_export_is_shareable(monkeypatch):
    from openpyxl import load_workbook

    page = {
        "snapshot": {
            "vendor_name": "SKF", "profile_code": "SKF",
            "planning_purpose": "replenishment", "snapshot_key": "SP-SKF",
            "as_of_date": "2026-09-01",
        }
    }
    report = {"uncovered": [{
        "internal_sku": "SKF-1", "active_months_24": 8,
        "sales_12": 12, "cali_inventory": 0, "cali_transit": 0,
        "bogota_inventory": 1, "suggested_quantity": 3,
        "assigned_to": "Jean Pierre", "reason": "Bogotá no puede cubrir.",
    }]}
    monkeypatch.setattr(
        StockPlanningRepository, "snapshot_detail", lambda _: page
    )
    monkeypatch.setattr(
        StockReplenishmentService, "report", lambda _: report
    )

    stream, filename = StockPlanningExportService.replenishment_uncovered(1)
    sheet = load_workbook(io.BytesIO(stream.read())).active

    assert filename.startswith("faltantes-cali-para-compras-")
    assert sheet["A6"].value == "SKF-1"
    assert sheet["G6"].value == 3
    assert sheet["H6"].value == "Jean Pierre"


def test_replenishment_transfer_export_excludes_pending_rows(monkeypatch):
    from openpyxl import load_workbook

    page = {
        "snapshot": {
            "vendor_name": "SKF", "planning_purpose": "replenishment",
            "snapshot_key": "SP-SKF", "as_of_date": "2026-09-01",
        },
        "products": [],
    }
    forecast = {
        "rows": [], "purchase_export_ready": False,
        "transfer_export_ready": False,
        "transfers": [
            {
                "sku": "APPROVED", "from_branch": "1", "to_branch": "50",
                "final_quantity": 2, "decision": {"decision_status": "approved"},
                "internal_move_16_to_1": 0,
            },
            {
                "sku": "PENDING", "from_branch": "1", "to_branch": "50",
                "final_quantity": 4, "decision": None,
                "internal_move_16_to_1": 0,
            },
        ],
    }
    monkeypatch.setattr(
        StockPlanningExportService, "_data", lambda _: (page, forecast)
    )

    stream, _ = StockPlanningExportService.transfers(1)
    book = load_workbook(io.BytesIO(stream.read()))
    sheet = book["Bogotá a Cali"]

    assert book.sheetnames == ["Bogotá a Cali"]
    assert sheet["A6"].value == "APPROVED"
    assert sheet["A7"].value is None
    assert sheet["H6"].value == "Aprobada manualmente"
