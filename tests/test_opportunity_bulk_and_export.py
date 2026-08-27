from datetime import date, timedelta
from unittest.mock import patch

from openpyxl import load_workbook

from app.workspace.services.opportunity_bulk_service import (
    OpportunityBulkUpdateService,
)
from app.workspace.services.opportunity_export_service import (
    OpportunityExportService,
)


def test_bulk_update_changes_open_stage_and_creates_next_action():
    future = (date.today() + timedelta(days=3)).isoformat()
    with (
        patch(
            "app.workspace.services.opportunity_bulk_service."
            "ProjectWorkspaceService.get_workspace",
            side_effect=lambda project_id: {
                "project": {"id": project_id, "status": "prospect"}
            },
        ),
        patch(
            "app.workspace.services.opportunity_bulk_service."
            "ProjectWorkspaceService.change_status"
        ) as change_status,
        patch(
            "app.workspace.services.opportunity_bulk_service."
            "FollowupRepository.find_pending_duplicate",
            return_value=None,
        ),
        patch(
            "app.workspace.services.opportunity_bulk_service."
            "FollowupRepository.create_followup"
        ) as create_followup,
    ):
        result = OpportunityBulkUpdateService.apply.__wrapped__(
            OpportunityBulkUpdateService,
            project_ids=[3, 4], new_status="quoting",
            followup_date=future,
            followup_description="Actualizar avance", actor="manager",
        )

    assert result == {
        "selected": 2, "changed_statuses": 2, "created_followups": 2,
    }
    assert change_status.call_count == 2
    assert create_followup.call_count == 2


def test_bulk_update_does_not_close_opportunities_without_individual_review():
    try:
        OpportunityBulkUpdateService.apply.__wrapped__(
            OpportunityBulkUpdateService,
            project_ids=[1], new_status="won"
        )
    except ValueError as error:
        assert "revisan individualmente" in str(error)
    else:
        raise AssertionError("A terminal bulk status must be rejected")


def test_filtered_export_is_a_valid_excel_workbook():
    item = {
        "id": 7, "origin_reference": "PC-7", "name": "Rodamiento",
        "customer_name": "Cliente", "sales_rep": "Ana", "office": "Cali",
        "status_label": "Cotización", "crm_status": "Abierto",
        "crm_stage": "Propuesta o Cotizacion",
        "crm_source_date": "2026-08-01", "crm_close_date": "2026-09-01",
        "commercial_amount": None, "quote": None,
        "crm_potential_value": 1250000, "health": {"label": "Activa"},
        "next_action_date": "2026-08-30", "current_blocker": None,
        "last_activity_at": "2026-08-20 10:00:00",
    }
    page = {
        "opportunities": [item],
        "filters": {"office": "Cali"},
    }
    with patch(
        "app.workspace.services.opportunity_export_service."
        "OpportunityListService.get_page", return_value=page,
    ):
        stream, filename = OpportunityExportService.build({"office": "Cali"})

    workbook = load_workbook(stream, data_only=False)
    sheet = workbook["Oportunidades"]
    assert filename.startswith("oportunidades-Cali-")
    assert sheet["A1"].value == "Pipeline de oportunidades"
    headers = [cell.value for cell in sheet[4]]
    assert "Vendedor" in headers
    assert "Próxima acción" in headers
    assert sheet.cell(row=5, column=headers.index("Valor comercial") + 1).value == 1250000
    assert sheet.cell(row=5, column=headers.index("Fuente del valor") + 1).value == "Potencial CRM · por revisar"
