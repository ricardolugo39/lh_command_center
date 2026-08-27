import io
from decimal import Decimal
from pathlib import Path

import pytest
import xlwt
from openpyxl import Workbook
from werkzeug.datastructures import FileStorage

from app.database.migrations import upgrade
from app.workspace.connectors.agreement_workbook_parser import (
    AgreementWorkbookError,
    AgreementWorkbookParser,
)
from app.workspace.services.agreement_import_service import (
    AgreementImportError,
    AgreementImportService,
)
from app.workspace.services.agreement_import_validator import AgreementImportValidator


def _workbook(headers=None, rows=None, second_sheet=False) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Productos"
    sheet.append(headers or ["Código", "Descripción", "Precio negociado", "Moneda"])
    for row in rows or [["ABC-1", "Rodamiento", 100, "COP"]]:
        sheet.append(row)
    if second_sheet:
        workbook.create_sheet("Notas")
    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _xls_workbook(headers=None, rows=None, second_sheet=False) -> bytes:
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet("Productos")
    headers = headers or ["Código", "Descripción", "Precio negociado", "Moneda"]
    rows = rows or [["ABC-1", "Rodamiento", 100, "COP"]]
    for column, value in enumerate(headers):
        sheet.write(0, column, value)
    for row_index, row in enumerate(rows, start=1):
        for column, value in enumerate(row):
            sheet.write(row_index, column, value)
    if second_sheet:
        workbook.add_sheet("Notas")
    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _html_workbook(*, multiple=False) -> bytes:
    small = "<table><tr><th>Ignorar</th></tr></table>"
    products = """
        <table><tr><th>Código</th><th>Descripción</th><th>Precio negociado</th><th>Moneda</th></tr>
        <tr><td>ABC-1</td><td>Rodamiento</td><td>100</td><td>COP</td></tr></table>
    """
    extra = small if multiple else ""
    return f"<!DOCTYPE html><html><script>alert('no')</script><body>{extra}{products}</body></html>".encode()


def _papeles_html_workbook(row_count=1040) -> bytes:
    metadata = """<table><tr><th>Tipo</th><td>AP (Acuerdo de Precios)</td></tr>
    <tr><th>Cliente</th><td>PAPELES NACIONALES S.A.</td></tr>
    <tr><th>Fecha Inicio</th><td>2026-01-01</td></tr>
    <tr><th>Fecha Fin</th><td>2026-12-31</td></tr></table>"""
    headers = """<table><thead><th>NUMERO PARTE</th><th>REFERENCIA SKF</th>
    <th>FOB DD LISTA</th><th>FOB DD CONVENIO</th><th>PRECIO SUGERIDO</th>
    <th>PRODUCT LINE</th><th>SPC</th></thead>"""
    rows = []
    for index in range(row_count):
        sku = "R024D62094" if index == 0 else f"P{index:05d}"
        reference = "628/4-2Z" if index == 0 else f"SKF-{index:05d}"
        rows.append(f"<tr><td>{sku}</td><td>{reference}</td><td>5, 72</td><td>4, 58</td><td>6, 91</td><td>DGBB</td><td></td></tr>")
    return (metadata + headers + "".join(rows) + "</table>").encode()


def test_parser_detects_aliases_sheets_blank_and_numeric_rows(tmp_path):
    path = tmp_path / "agreement.xlsx"
    path.write_bytes(_workbook(rows=[["ABC", "Producto", 25.5, "USD"], [None] * 4], second_sheet=True))

    parsed = AgreementWorkbookParser.inspect(path)

    assert parsed["worksheets"] == ["Productos", "Notas"]
    assert parsed["mapping"]["internal_sku"] == "Código"
    assert parsed["mapping"]["negotiated_price"] == "Precio negociado"
    assert parsed["rows"][0]["values"][2] == 25.5


def test_parser_rejects_corrupt_workbook(tmp_path):
    path = tmp_path / "bad.xlsx"
    path.write_bytes(b"not a workbook")
    with pytest.raises(AgreementWorkbookError):
        AgreementWorkbookParser.inspect(path)


def test_xls_and_xlsx_produce_the_same_normalized_model(tmp_path):
    xlsx = tmp_path / "agreement.xlsx"
    xls = tmp_path / "agreement.xls"
    xlsx.write_bytes(_workbook(second_sheet=True))
    xls.write_bytes(_xls_workbook(second_sheet=True))

    xlsx_model = AgreementWorkbookParser.inspect(xlsx)
    xls_model = AgreementWorkbookParser.inspect(xls)

    assert xls_model == xlsx_model


def test_html_xls_produces_the_same_parser_data_as_excel(tmp_path):
    xlsx = tmp_path / "agreement.xlsx"
    html_xls = tmp_path / "erp-export.xls"
    xlsx.write_bytes(_workbook())
    html_xls.write_bytes(_html_workbook())

    excel = AgreementWorkbookParser.inspect(xlsx)
    html = AgreementWorkbookParser.inspect(html_xls)

    assert html["headers"] == excel["headers"]
    assert html["mapping"] == excel["mapping"]
    assert html["rows"] == excel["rows"]


def test_html_xls_selects_largest_table(tmp_path):
    path = tmp_path / "multiple.xls"
    path.write_bytes(_html_workbook(multiple=True))
    parsed = AgreementWorkbookParser.inspect(path)
    assert parsed["selected_worksheet"] == "Tabla 2"
    assert parsed["headers"][0] == "Código"


def test_papeles_export_preserves_headers_rows_mapping_and_metadata(tmp_path):
    path = tmp_path / "papeles_nacionales.xls"
    path.write_bytes(_papeles_html_workbook())
    parsed = AgreementWorkbookParser.inspect(path)

    assert parsed["selected_worksheet"] == "Tabla 2"
    assert parsed["headers"] == [
        "NUMERO PARTE", "REFERENCIA SKF", "FOB DD LISTA",
        "FOB DD CONVENIO", "PRECIO SUGERIDO", "PRODUCT LINE", "SPC",
    ]
    assert len(parsed["rows"]) == 1040
    assert parsed["rows"][0]["values"][0:2] == ["R024D62094", "628/4-2Z"]
    assert parsed["mapping"]["internal_sku"] == "NUMERO PARTE"
    assert parsed["mapping"]["manufacturer_part_number"] == "REFERENCIA SKF"
    assert parsed["mapping"]["negotiated_price"] == "FOB DD CONVENIO"
    assert parsed["mapping"]["list_price"] == "FOB DD LISTA"
    assert parsed["detected_metadata"] == {
        "agreement_type": "AP (Acuerdo de Precios)",
        "customer_name": "PAPELES NACIONALES S.A.",
        "start_date": "2026-01-01", "end_date": "2026-12-31",
    }
    validation = AgreementImportValidator.validate(
        {"name": "Papeles", "supplier": "SKF", "currency": "USD", "start_date": "2026-01-01", "end_date": "2026-12-31"},
        parsed, parsed["mapping"],
    )
    assert validation["summary"]["total"] == 1040
    assert validation["summary"]["blank"] == 0
    assert validation["summary"]["valid"] == 1040
    assert validation["summary"]["warnings"] == 0
    assert validation["rows"][0]["negotiated_price"] == Decimal("4.58")
    assert validation["rows"][0]["list_price"] == Decimal("5.72")
    assert validation["rows"][0]["suggested_price"] == Decimal("6.91")
    contradicted = AgreementImportValidator.validate(
        {"name": "Papeles", "supplier": "SKF", "currency": "USD", "start_date": "2025-01-01", "end_date": "2025-12-31"},
        parsed, parsed["mapping"],
    )
    assert "vigencia diferente" in contradicted["metadata_warnings"][0]


@pytest.mark.parametrize(
    "content",
    [b"<html><body>Sin datos</body></html>", b"<html><table><tr><td>incompleto"],
)
def test_html_xls_without_valid_table_is_rejected(tmp_path, content):
    path = tmp_path / "empty-export.xls"
    path.write_bytes(content)
    with pytest.raises(AgreementWorkbookError, match="tabla de datos utilizable"):
        AgreementWorkbookParser.inspect(path)


def test_parser_rejects_corrupt_xls(tmp_path):
    path = tmp_path / "bad.xls"
    path.write_bytes(b"not an xls workbook")
    with pytest.raises(AgreementWorkbookError, match="no contiene"):
        AgreementWorkbookParser.inspect(path)


def test_reader_uses_content_instead_of_extension(tmp_path):
    path = tmp_path / "openxml-export.xls"
    path.write_bytes(_workbook())
    parsed = AgreementWorkbookParser.inspect(path)
    assert parsed["mapping"]["internal_sku"] == "Código"


@pytest.mark.parametrize("extension", [".xls", ".xlsx"])
def test_parser_rejects_empty_workbook(tmp_path, extension):
    content = (
        _xls_workbook(headers=[""], rows=[])
        if extension == ".xls"
        else _workbook(headers=[""], rows=[])
    )
    path = tmp_path / f"empty{extension}"
    path.write_bytes(content)
    with pytest.raises(AgreementWorkbookError, match="vacía"):
        AgreementWorkbookParser.inspect(path)


def test_validator_reports_warnings_errors_duplicates_and_blanks():
    parsed = {
        "headers": ["Código", "Precio"],
        "rows": [
            {"source_row_number": 2, "values": ["A", 10]},
            {"source_row_number": 3, "values": ["A", 10]},
            {"source_row_number": 4, "values": ["B", -1]},
            {"source_row_number": 5, "values": [None, None]},
        ],
    }
    result = AgreementImportValidator.validate(
        {"name": "Acuerdo", "supplier": "SKF", "currency": "USD", "start_date": "2026-01-01", "end_date": "2026-12-31"},
        parsed,
        {"internal_sku": "Código", "negotiated_price": "Precio"},
    )
    assert result["summary"]["duplicates"] == 1
    assert result["summary"]["blank"] == 1
    assert result["summary"]["errors"] == 1
    assert result["summary"]["warnings"] == 1
    assert result["can_confirm"] is False


def test_decimal_price_normalizes_spaces_and_colombian_grouping():
    assert AgreementImportValidator._decimal_price("5, 72") == Decimal("5.72")
    assert AgreementImportValidator._decimal_price("1.098, 99") == Decimal("1098.99")


def test_product_mapping_excludes_agreement_level_fields():
    assert "currency" not in AgreementImportValidator.DESTINATIONS
    assert "product_start_date" not in AgreementImportValidator.DESTINATIONS
    assert "product_end_date" not in AgreementImportValidator.DESTINATIONS


def test_agreement_currency_is_required_without_product_currency():
    parsed = {
        "headers": ["SKU"],
        "rows": [{"source_row_number": 2, "values": ["ABC"]}],
    }
    result = AgreementImportValidator.validate(
        {"name": "Acuerdo", "supplier": "SKF", "currency": "",
         "start_date": "2026-01-01", "end_date": "2026-12-31"},
        parsed, {"internal_sku": "SKU"},
    )
    assert "El campo moneda es obligatorio." in result["blocking_errors"]


@pytest.fixture
def import_environment(tmp_path, monkeypatch):
    database = tmp_path / "database.db"
    monkeypatch.setattr("app.database.connection.DB_PATH", database)
    monkeypatch.setattr("app.workspace.services.agreement_import_service.STAGING_ROOT", tmp_path / "staging")
    monkeypatch.setattr("app.workspace.services.agreement_import_service.AGREEMENT_ROOT", tmp_path / "agreements")
    upgrade()
    from app.database.connection import get_connection
    with get_connection() as connection:
        connection.execute("INSERT INTO ws_customers (name, erp_customer_id) VALUES ('Cliente', '1')")
        connection.commit()
    return tmp_path


def _stage(extension: str = ".xlsx") -> FileStorage:
    content = _xls_workbook() if extension == ".xls" else _workbook()
    mime = "application/vnd.ms-excel" if extension == ".xls" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return FileStorage(stream=io.BytesIO(content),
                       filename=f"Acuerdo real{extension}", content_type=mime)


def test_confirmation_is_atomic_and_preserves_document(import_environment):
    token = AgreementImportService.stage(1, _stage(), {
        "name": "Acuerdo 2026", "supplier": "SKF", "currency": "USD",
        "start_date": "2026-01-01", "end_date": "2026-12-31", "notes": "",
    })
    agreement_id = AgreementImportService.confirm(1, token, replace_active=False)
    from app.database.connection import get_connection
    with get_connection() as connection:
        assert connection.execute("SELECT COUNT(*) FROM ws_agreements").fetchone()[0] == 1
        assert connection.execute("SELECT COUNT(*) FROM ws_agreement_items WHERE agreement_id = ?", (agreement_id,)).fetchone()[0] == 1
        document = connection.execute("SELECT stored_name FROM ws_agreement_documents").fetchone()
    assert (import_environment / "agreements" / str(agreement_id) / document[0]).exists()
    with pytest.raises(AgreementImportError):
        AgreementImportService.confirm(1, token, replace_active=False)


def test_xls_confirmation_preserves_original_format_and_metadata(import_environment):
    original_content = _xls_workbook()
    uploaded_file = FileStorage(
        stream=io.BytesIO(original_content), filename="Acuerdo real.xls",
        content_type="application/vnd.ms-excel",
    )
    token = AgreementImportService.stage(1, uploaded_file, {
        "name": "Acuerdo XLS", "supplier": "SKF", "currency": "USD",
        "start_date": "2026-01-01", "end_date": "2026-12-31", "notes": "",
    })
    agreement_id = AgreementImportService.confirm(1, token, replace_active=False)
    from app.database.connection import get_connection
    with get_connection() as connection:
        document = connection.execute("""
            SELECT original_name, stored_name, file_extension, mime_type,
                   file_size, created_at
            FROM ws_agreement_documents WHERE agreement_id = ?
        """, (agreement_id,)).fetchone()
        item = connection.execute("""
            SELECT internal_sku, description, negotiated_price_decimal, price_currency
            FROM ws_agreement_items WHERE agreement_id = ?
        """, (agreement_id,)).fetchone()
    assert document[0].endswith(".xls")
    assert document[1].endswith(".xls")
    assert document[2] == ".xls"
    assert document[3] == "application/vnd.ms-excel"
    assert document[4] > 0 and document[5]
    assert tuple(item) == ("ABC-1", "Rodamiento", "100.0", None)
    with get_connection() as connection:
        assert connection.execute(
            "SELECT currency FROM ws_agreements WHERE id = ?", (agreement_id,)
        ).fetchone()[0] == "USD"
    stored_path = (
        import_environment / "agreements" / str(agreement_id) / document[1]
    )
    assert stored_path.read_bytes() == original_content


def test_html_xls_uses_same_preview_and_confirmation_pipeline(import_environment):
    content = _html_workbook()
    uploaded = FileStorage(
        stream=io.BytesIO(content), filename="Exportacion ERP.xls",
        content_type="application/vnd.ms-excel",
    )
    token = AgreementImportService.stage(1, uploaded, {
        "name": "Acuerdo ERP", "supplier": "Fabricante", "currency": "USD",
        "start_date": "2026-01-01", "end_date": "2026-12-31", "notes": "",
    })
    preview = AgreementImportService.preview(1, token)
    assert preview["validation"]["can_confirm"] is True
    assert preview["validation"]["rows"][0]["internal_sku"] == "ABC-1"
    agreement_id = AgreementImportService.confirm(1, token, replace_active=False)
    from app.database.connection import get_connection
    with get_connection() as connection:
        item = connection.execute(
            "SELECT internal_sku, negotiated_price_decimal FROM ws_agreement_items WHERE agreement_id = ?",
            (agreement_id,),
        ).fetchone()
    assert tuple(item) == ("ABC-1", "100")


def test_product_failure_rolls_back_agreement_and_file(import_environment, monkeypatch):
    token = AgreementImportService.stage(1, _stage(), {
        "name": "Acuerdo", "supplier": "SKF", "currency": "USD", "start_date": "2026-01-01",
        "end_date": "2026-12-31", "notes": "",
    })
    monkeypatch.setattr(
        "app.workspace.services.agreement_import_service.AgreementItemRepository.insert_imported_items",
        lambda *args, **kwargs: (_ for _ in ()).throw(RuntimeError("failure")),
    )
    with pytest.raises(RuntimeError):
        AgreementImportService.confirm(1, token, replace_active=False)
    from app.database.connection import get_connection
    with get_connection() as connection:
        assert connection.execute("SELECT COUNT(*) FROM ws_agreements").fetchone()[0] == 0
        assert connection.execute("SELECT COUNT(*) FROM ws_agreement_documents").fetchone()[0] == 0
    assert not list((import_environment / "agreements").rglob("*.xlsx"))


def test_existing_active_agreement_requires_explicit_replacement(import_environment):
    from app.database.connection import get_connection
    with get_connection() as connection:
        connection.execute("INSERT INTO ws_agreements (customer_id, name, status) VALUES (1, 'Anterior', 'active')")
        connection.commit()
    token = AgreementImportService.stage(1, _stage(), {
        "name": "Nuevo", "supplier": "SKF", "currency": "USD", "start_date": "2026-01-01",
        "end_date": "2026-12-31", "notes": "",
    })
    with pytest.raises(AgreementImportError, match="expresamente"):
        AgreementImportService.confirm(1, token, replace_active=False)
    new_id = AgreementImportService.confirm(1, token, replace_active=True)
    with get_connection() as connection:
        statuses = connection.execute("SELECT id, status FROM ws_agreements ORDER BY id").fetchall()
    assert [(row[0], row[1]) for row in statuses] == [(1, "expired"), (new_id, "active")]


def test_legacy_skf_and_generic_items_share_detail_read_model(import_environment):
    from app.database.connection import get_connection
    from app.workspace.repositories.agreement_item_repository import AgreementItemRepository
    with get_connection() as connection:
        connection.execute("INSERT INTO ws_agreements (customer_id, name, status) VALUES (1, 'Legado', 'active')")
        connection.execute("""
            INSERT INTO ws_agreement_items (
                agreement_id, part_number, skf_reference, agreement_price_usd
            ) VALUES (1, 'LEG-1', 'SKF-1', 25)
        """)
        connection.commit()
    items = AgreementItemRepository.list_imported_items(1)
    assert items[0]["part_number"] == "LEG-1"
    assert items[0]["internal_sku"] is None


def test_existing_agreement_service_operations_commit_at_service_boundary(import_environment):
    from app.workspace.services.agreement_service import AgreementService
    agreement_id = AgreementService.create(
        customer_id=1, agreement_number="", name="Manual", status="draft",
        agreement_type="", supplier="SKF", annual_target=None, currency="COP",
        start_date=None, end_date=None, renewal_date=None,
        has_consignment=False, notes="",
    )
    assert AgreementService.get(agreement_id)["name"] == "Manual"
