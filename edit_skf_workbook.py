import re
import sqlite3
from collections import defaultdict
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter


SOURCE = Path("Negociación SKF 2026.xlsx")
OUTPUT = Path("outputs/skf_homologacion/Negociación SKF 2026 homologada DB Sales.xlsx")
DB_PATH = Path("database/commercial.db")


def clean_ref(value: str) -> str:
    value = re.sub(r"\s+", " ", value.strip().upper())
    value = re.sub(r"\s*/\s*", "/", value)
    return value


def candidate(description: str) -> tuple[str, bool]:
    d = clean_ref(description)
    explicit_skf = "SKF" in d

    patterns = [
        r"\b(45[135]-\d{3}-\d{3}-VS)\b",
        r"\b(LG(?:MT|HB)\s*\d(?:/\d+)?)\b",
        r"\b(PCMF?\s*\d+(?:\.\d+)?\s*E)\b",
        r"\b(PHF\s+TB\d+X\d+MM)\b",
        r"\b(PHE\s+L\d+(?:HUB|NR))\b",
        r"\b(FRB\s*\d+/\d+)\b",
        r"\b(ASNH\s*\d+-\d+)\b",
        r"\b(FSAF\s*\d+)\b",
        r"\b(SNL\s*\d+-\d+)\b",
        r"\b(HA?\s*\d{3})\b",
        r"\b(SY\s*\d+(?:\.\d+/\d+)?\s*TF)\b",
        r"\b(FY\s*\d+\s*(?:FM|TF))\b",
        r"\b(Y(?:AR|AT|ET)\s*\d+(?:\s*\d+)?(?:\s*2F)?)\b",
        r"\b(LBCR\s*\d+\s*A\s*2LS(?:\s*CM\d+)?)\b",
        r"\b(KR\s*\d+\s*PPA)\b",
        r"\b(SIL?KB\s*\d+\s*(?:ES?|F))\b",
        r"\b((?:SIL?|SAL?)\s*\d+\s*E(?:S)?)\b",
        r"\b((?:KM|MB)\s*\d+)\b",
        r"\b(W\s*\d{3,5}\s*2RS1)\b",
        r"\b(NU\s*\d+\s*(?:ECP|ECJ)?(?:\s*C\d)?)\b",
        r"\b(NJ\s*\d+\s*ECP(?:\s*C\d)?)\b",
        r"\b(NA\s*\d+\s*2RS)\b",
        r"\b(HK\s*\d+)\b",
        r"\b(\d{4,5}\s*(?:BECBP|EKC3|EK\s*C3|CCK\s*C3|K/C3|ECP|ECJ))\b",
        r"\b(\d{3,5}\s*(?:2RSH/C3|2RS1/C3|2RSH|2RS1|2ZNR|2Z/C3|2Z\s*C3|2Z|RS|C3))\b",
        r"\b(\d{3,5}\s*(?:2RSR|2RS(?:\s*C3)?|2RS\s*SS))\b",
        r"\b(\d{4,5})\b",
    ]
    for pattern in patterns:
        match = re.search(pattern, d)
        if match:
            ref = clean_ref(match.group(1))
            ref = re.sub(r"\s+(2RS1|2RSH|2ZNR|2Z|RS)\s+(C3)$", r" \1/\2", ref)
            ref = re.sub(r"^(\d{4})BECBP$", r"\1 BECBP", ref)
            ambiguous = bool(re.search(r"(?:^|\s)2RS(?:\s|$)|2RSR|\bRS$", d)) and not re.search(r"2RS1|2RSH", d)
            certain = explicit_skf and not ambiguous
            if re.search(r"BECBP|ECP|ECJ|EKC3|2RS1|2RSH|2ZNR|2Z(?:/C3)?$|K/C3|PPA", ref):
                certain = not ("IKO" in d or "INA" in d)
            if re.match(r"^(?:LG|PCM|PCMF|PHF|PHE|FRB|ASNH|FSAF|SNL|HA?\s|SY\s|FY\s|YAR\s|YAT\s|YET\s|LBCR\s|KR\s|SI|SIL|SA|SAL|NU\s|NJ\s|NA\s|45[135]-)", ref):
                certain = not ("IKO" in d or "INA" in d)
            return ref, certain

    dims = re.search(r"\b(\d+(?:\.\d+)?\s*[Xx]\s*\d+(?:\.\d+)?(?:\s*[Xx]\s*\d+(?:/\d+)?)?\s*MM)\b", description, re.I)
    if dims:
        return re.sub(r"[Xx]", " x ", clean_ref(dims.group(1))), False

    generic_patterns = [
        r"\b(417515)\b", r"\b(5MR625-35)\b", r"\b(SMR900-25)\b",
        r"\b(BX\s*82)\b", r"\b(XPB\s*\d+)\b", r"\b(300\s*H\s*100)\b",
        r"\b(LME\d+NUU)\b", r"\b(KBS\s*\d+PP)\b", r"\b(HGH\d+(?:CA|HA))\b",
        r"\b(F4B-SC-108)\b", r"\b(E5)\b", r"\b(5JE)\b", r"\b(L-?\d+)\b",
    ]
    for pattern in generic_patterns:
        match = re.search(pattern, d)
        if match:
            return clean_ref(match.group(1)), False

    # Last-resort candidate: keep the description so no row is silently left unresolved.
    return description.strip(), False


def result_for(description):
    if not description:
        return None
    ref, certain = candidate(str(description))
    return ref if certain else f"Investigar — {ref}"


def db_key(value: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", str(value).upper())


def load_sales_references():
    uri = f"file:{DB_PATH.resolve()}?mode=ro"
    with sqlite3.connect(uri, uri=True) as conn:
        rows = conn.execute(
            """
            SELECT TRIM(prefijo_1), UPPER(TRIM(COALESCE(sufijo,''))), COUNT(*), MAX(fecha)
            FROM raw_sales
            WHERE TRIM(COALESCE(prefijo_1,'')) <> ''
            GROUP BY TRIM(prefijo_1), UPPER(TRIM(COALESCE(sufijo,'')))
            """
        ).fetchall()
    by_key = defaultdict(list)
    all_refs = []
    for ref, brand, uses, latest in rows:
        item = {"ref": ref, "brand": brand, "uses": uses, "latest": latest or ""}
        by_key[db_key(ref)].append(item)
        all_refs.append((db_key(ref), item))
    return by_key, all_refs


DB_BY_KEY, DB_REFS = load_sales_references()


def choose(items, wants_skf=False):
    pool = [x for x in items if x["brand"] == "SKF"] if wants_skf else list(items)
    if not pool:
        pool = list(items)
    return sorted(pool, key=lambda x: (x["uses"], x["latest"]), reverse=True)[0]


def bearing_aliases(raw_candidate: str):
    text = clean_ref(raw_candidate)
    match = re.search(r"\b(608|\d{4,5})\s*(?:-|\s)?(2RS|2RSR|2RSH|2RS1|2Z|RS)?(?:/|\s)?(C3)?\b", text)
    if not match:
        return []
    size, seal, play = match.groups()
    if not seal:
        return []
    if seal == "2RS":
        seal = "2RSH" if size == "608" or int(size[-2:]) <= 5 else "2RS1"
    base = f"{size}-{seal}"
    if play:
        base += "/C3"
    return [base, base + "GJN"]


def match_sales_db(description: str):
    preliminary = result_for(description)
    candidate_text = preliminary.split("—", 1)[-1].strip()
    wants_skf = "SKF" in clean_ref(description)

    exact = DB_BY_KEY.get(db_key(candidate_text), [])
    if exact:
        picked = choose(exact, wants_skf=wants_skf)
        return picked["ref"], "exacto"

    candidate_key = db_key(candidate_text)
    if len(candidate_key) >= 5:
        expanded = [
            item for key, item in DB_REFS
            if candidate_key in key and (not wants_skf or item["brand"] == "SKF")
        ]
        if expanded:
            unique_refs = {item["ref"] for item in expanded}
            picked = choose(expanded, wants_skf=wants_skf)
            if len(unique_refs) == 1:
                return picked["ref"], "expandido_db"
            return f"Investigar — {candidate_text}", "ambiguo_db"

    for alias in bearing_aliases(candidate_text):
        matches = DB_BY_KEY.get(db_key(alias), [])
        if matches:
            picked = choose(matches, wants_skf=True)
            return picked["ref"], "alias_rodamiento"

    description_key = db_key(description)
    contained = defaultdict(list)
    for key, item in DB_REFS:
        if len(key) >= 5 and re.search(r"[A-Z]", key) and re.search(r"\d", key) and key in description_key:
            if not wants_skf or item["brand"] == "SKF":
                contained[len(key)].append(item)
    if contained:
        longest = max(contained)
        unique_refs = {item["ref"] for item in contained[longest]}
        picked = choose(contained[longest], wants_skf=wants_skf)
        if len(unique_refs) == 1:
            return picked["ref"], "contenido"
        return f"Investigar — {candidate_text}", "ambiguo_db"

    return f"Investigar — {candidate_text}", "sin_match"


wb = openpyxl.load_workbook(SOURCE)
audit = defaultdict(int)
for ws in wb.worksheets:
    if ws.title == "PURINA":
        desc_col, target_col = 6, 9
        style_col = 8
    else:
        desc_col, target_col = 2, 5
        style_col = 5  # The workbook already contains a formatted blank target column.

    if ws.title == "PURINA":
        for row in range(1, ws.max_row + 1):
            src, dst = ws.cell(row, style_col), ws.cell(row, target_col)
            dst._style = copy(src._style)
            if src.has_style:
                dst.font = copy(src.font)
                dst.fill = copy(src.fill)
                dst.border = copy(src.border)
                dst.alignment = copy(src.alignment)
                dst.protection = copy(src.protection)
                dst.number_format = src.number_format

    ws.cell(1, target_col).value = "Referencia LH"
    ws.cell(1, target_col).alignment = copy(ws.cell(1, desc_col).alignment)
    ws.cell(1, target_col).font = copy(ws.cell(1, desc_col).font)
    ws.cell(1, target_col).fill = copy(ws.cell(1, desc_col).fill)
    ws.cell(1, target_col).border = copy(ws.cell(1, desc_col).border)
    ws.cell(1, target_col).number_format = "@"

    for row in range(2, ws.max_row + 1):
        desc = ws.cell(row, desc_col).value
        if desc:
            matched, status = match_sales_db(str(desc))
            ws.cell(row, target_col).value = matched
            audit[status] += 1
            ws.cell(row, target_col).number_format = "@"
            ws.cell(row, target_col).alignment = copy(ws.cell(row, desc_col).alignment)

    letter = get_column_letter(target_col)
    ws.column_dimensions[letter].width = 34

OUTPUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUTPUT)
print(dict(sorted(audit.items())))
print(OUTPUT.resolve())
